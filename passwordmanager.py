from openpyxl import *
import random
import time


'''
throwing entire workbook:

for row in range(1,14):
	for col in range(0,3,2):
		char = chr(65 + col)
		print(ws[char + str(row)].value)	
	print()

'''
print()
username=input("Enter username : ")
password=input("Enter password : ")

if username=='xxx' and password=='xxx':

	sudo_password = 'xxx'
	otp = random.randint(1000,9999)
	print()
	print()
	print("Session Started! ")
	print()
	time.sleep(2)
	print("Your 'one time password' is",otp)
	print()
	print("It will be required while making any proccess after.")
	print()
	time.sleep(3)
	
	
	
	while True:

		print('''

			commands       - action

			get 	 	   - get password
			gethasher      - get decrypter
			exit           - exit out

			''')

		command=input("Enter command : ")
		if command.lower()=='get':
			otpconform=int(input("Enter otp provided : "))
			if otpconform==otp:
				emailask=input("Enter email id password required : ")

				wb=load_workbook('rawpasses.xlsx')
				ws=wb.active

				for row in range(2,14):
					patch=ws[chr(65)+str(row)].value
					if patch==emailask:
						sudoask=input("Enter sudo password : ")
						if sudoask==sudo_password:
							print()
							print("Hashed password : ")
							print(ws[chr(67)+str(row)].value)
							print()
						else:
							print("Wrong sudo password! ")
							break
			else:
				print("Wrong otp provided!")
				break
			
		if command.lower()=='gethasher':
			otpconform=int(input("Enter otp provided : "))
			if otpconform==otp:
				print()
				print("https://www.online-toolz.com/tools/text-encryption-decryption.php")
			else:
				print()
				print("Wrong otp provided!")
				break

		if command.lower()=='exit':
			break
else:
	print()
	print("Wrong credentials! ")
	print()



